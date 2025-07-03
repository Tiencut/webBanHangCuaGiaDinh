# Script Python để tạo file Excel mẫu cho test import
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import os

# Tạo workbook
wb = Workbook()

# Xóa sheet mặc định
default_sheet = wb.active
wb.remove(default_sheet)

# Tạo sheet Categories
categories_data = [
    {'name': 'Phụ tùng động cơ', 'description': 'Các phụ tùng liên quan đến động cơ xe tải', 'parentCategory': ''},
    {'name': 'Piston', 'description': 'Piston động cơ các loại', 'parentCategory': 'Phụ tùng động cơ'},
    {'name': 'Phụ tùng gầm xe', 'description': 'Các phụ tùng liên quan đến gầm xe', 'parentCategory': ''},
    {'name': 'Phanh', 'description': 'Hệ thống phanh', 'parentCategory': 'Phụ tùng gầm xe'}
]

ws_categories = wb.create_sheet("Categories")
# Header
headers = ['name', 'description', 'parentCategory']
for col, header in enumerate(headers, 1):
    cell = ws_categories.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

# Data
for row, data in enumerate(categories_data, 2):
    ws_categories.cell(row=row, column=1, value=data['name'])
    ws_categories.cell(row=row, column=2, value=data['description'])
    ws_categories.cell(row=row, column=3, value=data['parentCategory'])

# Tạo sheet Suppliers
suppliers_data = [
    {'name': 'Công ty TNHH Phụ Tùng ABC', 'contactInfo': 'Mr. Nguyễn Văn A - 0901234567', 'address': '123 Đường Lê Lợi, Q1, TP.HCM', 'email': 'contact@phutungabc.com', 'phone': '0901234567'},
    {'name': 'Công ty Phụ Tùng Việt Nam', 'contactInfo': 'Ms. Trần Thị B - 0912345678', 'address': '456 Đường Nguyễn Huệ, Q1, TP.HCM', 'email': 'info@phutungvn.com', 'phone': '0912345678'}
]

ws_suppliers = wb.create_sheet("Suppliers")
# Header
headers = ['name', 'contactInfo', 'address', 'email', 'phone']
for col, header in enumerate(headers, 1):
    cell = ws_suppliers.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

# Data
for row, data in enumerate(suppliers_data, 2):
    ws_suppliers.cell(row=row, column=1, value=data['name'])
    ws_suppliers.cell(row=row, column=2, value=data['contactInfo'])
    ws_suppliers.cell(row=row, column=3, value=data['address'])
    ws_suppliers.cell(row=row, column=4, value=data['email'])
    ws_suppliers.cell(row=row, column=5, value=data['phone'])

# Tạo sheet VehicleModels
vehicles_data = [
    {'brand': 'Hyundai', 'model': 'HD120SL', 'year': '2020', 'engineType': 'D4DB', 'description': 'Xe tải 8 tấn Hyundai HD120SL'},
    {'brand': 'Isuzu', 'model': 'NPR85K', 'year': '2021', 'engineType': '4JJ1', 'description': 'Xe tải 3.5 tấn Isuzu NPR85K'},
    {'brand': 'Hino', 'model': 'XZU720L', 'year': '2019', 'engineType': 'N04C', 'description': 'Xe tải 5 tấn Hino XZU720L'}
]

ws_vehicles = wb.create_sheet("VehicleModels")
# Header
headers = ['brand', 'model', 'year', 'engineType', 'description']
for col, header in enumerate(headers, 1):
    cell = ws_vehicles.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

# Data
for row, data in enumerate(vehicles_data, 2):
    ws_vehicles.cell(row=row, column=1, value=data['brand'])
    ws_vehicles.cell(row=row, column=2, value=data['model'])
    ws_vehicles.cell(row=row, column=3, value=data['year'])
    ws_vehicles.cell(row=row, column=4, value=data['engineType'])
    ws_vehicles.cell(row=row, column=5, value=data['description'])

# Tạo sheet Products
products_data = [
    {'name': 'Piston Hyundai HD120SL', 'sku': 'PST-HD120-001', 'price': '1500000', 'category': 'Piston', 'description': 'Piston động cơ D4DB cho xe Hyundai HD120SL', 'compatibleVehicles': 'Hyundai HD120SL', 'supplier': 'Công ty TNHH Phụ Tùng ABC'},
    {'name': 'Má phanh Isuzu NPR85K', 'sku': 'BRK-NPR85-001', 'price': '850000', 'category': 'Phanh', 'description': 'Má phanh trước cho xe Isuzu NPR85K', 'compatibleVehicles': 'Isuzu NPR85K', 'supplier': 'Công ty Phụ Tùng Việt Nam'}
]

ws_products = wb.create_sheet("Products")
# Header
headers = ['name', 'sku', 'price', 'category', 'description', 'compatibleVehicles', 'supplier']
for col, header in enumerate(headers, 1):
    cell = ws_products.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

# Data
for row, data in enumerate(products_data, 2):
    ws_products.cell(row=row, column=1, value=data['name'])
    ws_products.cell(row=row, column=2, value=data['sku'])
    ws_products.cell(row=row, column=3, value=data['price'])
    ws_products.cell(row=row, column=4, value=data['category'])
    ws_products.cell(row=row, column=5, value=data['description'])
    ws_products.cell(row=row, column=6, value=data['compatibleVehicles'])
    ws_products.cell(row=row, column=7, value=data['supplier'])

# Tạo sheet Customers
customers_data = [
    {'name': 'Công ty Vận Tải XYZ', 'phone': '0912345678', 'email': 'info@vantaixyz.com', 'address': '456 Đường Nguyễn Trãi, Q5, TP.HCM', 'customerType': 'BUSINESS', 'taxCode': '0123456789'},
    {'name': 'Anh Nguyễn Văn B', 'phone': '0987654321', 'email': 'nguyenvanb@gmail.com', 'address': '789 Đường Lê Văn Việt, Q9, TP.HCM', 'customerType': 'INDIVIDUAL', 'taxCode': ''}
]

ws_customers = wb.create_sheet("Customers")
# Header
headers = ['name', 'phone', 'email', 'address', 'customerType', 'taxCode']
for col, header in enumerate(headers, 1):
    cell = ws_customers.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

# Data
for row, data in enumerate(customers_data, 2):
    ws_customers.cell(row=row, column=1, value=data['name'])
    ws_customers.cell(row=row, column=2, value=data['phone'])
    ws_customers.cell(row=row, column=3, value=data['email'])
    ws_customers.cell(row=row, column=4, value=data['address'])
    ws_customers.cell(row=row, column=5, value=data['customerType'])
    ws_customers.cell(row=row, column=6, value=data['taxCode'])

# Tạo sheet TrainingContent
training_data = [
    {'title': 'Cách nhận diện Piston Hyundai', 'content': 'Piston Hyundai HD120SL có đặc điểm: đường kính 104mm, có 4 rãnh piston, chất liệu hợp kim nhôm. Mã số in trên đỉnh piston: D4DB-104.', 'type': 'IDENTIFICATION', 'tags': 'piston,hyundai,hd120sl,động cơ', 'difficulty': 'BASIC', 'productSku': 'PST-HD120-001'},
    {'title': 'Kiểm tra má phanh Isuzu', 'content': 'Má phanh Isuzu NPR85K cần kiểm tra độ dày tối thiểu 3mm. Bề mặt không được có vết nứt, mòn không đều. Màu sắc đều, không có vết cháy xém.', 'type': 'INSPECTION', 'tags': 'má phanh,isuzu,npr85k,an toàn', 'difficulty': 'INTERMEDIATE', 'productSku': 'BRK-NPR85-001'}
]

ws_training = wb.create_sheet("TrainingContent")
# Header
headers = ['title', 'content', 'type', 'tags', 'difficulty', 'productSku']
for col, header in enumerate(headers, 1):
    cell = ws_training.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

# Data
for row, data in enumerate(training_data, 2):
    ws_training.cell(row=row, column=1, value=data['title'])
    ws_training.cell(row=row, column=2, value=data['content'])
    ws_training.cell(row=row, column=3, value=data['type'])
    ws_training.cell(row=row, column=4, value=data['tags'])
    ws_training.cell(row=row, column=5, value=data['difficulty'])
    ws_training.cell(row=row, column=6, value=data['productSku'])

# Auto adjust column width
for ws in wb.worksheets:
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

# Lưu file
output_path = '/workspaces/webBanHangCuaGiaDinh/sample-import-data.xlsx'
wb.save(output_path)

print(f"✅ Đã tạo file Excel mẫu: {output_path}")
print(f"📋 File chứa {len(wb.worksheets)} sheets:")
for i, ws in enumerate(wb.worksheets, 1):
    print(f"   {i}. {ws.title} ({ws.max_row-1} dòng dữ liệu)")
